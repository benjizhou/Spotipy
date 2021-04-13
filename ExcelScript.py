import CurrentlyPlaying
from datetime import datetime
from openpyxl import load_workbook
from os import startfile

def run():
    # Workbook() takes one, non-optional, argument  
    # which is the filename that we want to create. 
    workbook = load_workbook('Songs.xlsx')
    
    # The workbook object is then used to add new  
    # worksheet via the add_worksheet() method. 
    worksheet = workbook.active
    
    # Use the worksheet object to write 
    # data via the write() method.
    count = worksheet['Z1'].value
    a = 'A' + str(count)
    b = 'B' + str(count)
    c = 'C' + str(count)

    b1 = 'B' + str(count - 1)
    c1 = 'C' + str(count - 1)

    if (CurrentlyPlaying.playing() != None):
        
        name = CurrentlyPlaying.name()
        artist = CurrentlyPlaying.artist()

        if (name != worksheet[b1].value or artist != worksheet[c1].value):

            worksheet[a] = str(datetime.now())
            worksheet[b] = name
            worksheet[c] = artist
        
            worksheet['Z1'] = count + 1

            print(CurrentlyPlaying.name())
            # print(CurrentlyPlaying.rec())
            # print(CurrentlyPlaying.features(CurrentlyPlaying.id))
            # Finally, close the Excel file 
            # via the close() method. 
            workbook.save('Songs.xlsx')