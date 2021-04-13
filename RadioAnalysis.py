import sys
import json
import spotipy
import spotipy.util as util
from openpyxl import load_workbook
from os import startfile

scope = 'user-read-currently-playing'
username = 'benjizhou'
token = util.prompt_for_user_token(username, scope)

spotify = spotipy.Spotify(auth=token)

# Will create the playlist

# Will add each song to the Excel

# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
workbook = load_workbook('Analysis.xlsx')

# The workbook object is then used to add new
# worksheet via the add_worksheet() method.
worksheet = workbook.active

# print(json.dumps(spotify.audio_features(current['item']['uri']), indent=4))

URIs = spotify.current_user_recently_played(limit=50, after=None, before=None)['items']
count = 2
for i in URIs:
    uri = i['track']['uri']
    track = spotify.audio_features(uri)
    print(json.dumps(track))
    worksheet['A' + str(count)] = count
    worksheet['B' + str(count)] = uri
    worksheet['C' + str(count)] = i['track']['name']
    worksheet['D' + str(count)] = track[0]['danceability']
    worksheet['E' + str(count)] = track[0]['energy']
    worksheet['F' + str(count)] = track[0]['key']
    worksheet['G' + str(count)] = track[0]['loudness']
    worksheet['H' + str(count)] = track[0]['mode']
    worksheet['I' + str(count)] = track[0]['speechiness']
    worksheet['J' + str(count)] = track[0]['acousticness']
    worksheet['K' + str(count)] = track[0]['instrumentalness']
    worksheet['L' + str(count)] = track[0]['liveness']
    worksheet['M' + str(count)] = track[0]['valence']
    worksheet['N' + str(count)] = track[0]['tempo']
    worksheet['O' + str(count)] = int(track[0]['duration_ms']) / 1000
    count += 1
# print(spotify.audio_analysis(current['item']['uri']))

workbook.save('Analysis.xlsx')
workbook.close()